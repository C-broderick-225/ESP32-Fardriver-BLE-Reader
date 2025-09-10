import struct
import sys
import argparse


# Constants for parsing
EXPECTED_ID = 0x18F88006
PACKET_LENGTH = 12

def parse_sabvoton_packet(packet_bytes):
    """Parses a 12-byte serial packet structured like a CAN frame."""
    
    # Unpack the 12 bytes into the ID and the 8-byte payload
    # ">I" = Big-Endian 4-byte integer (for the ID)
    # "8s" = 8-byte string/bytes (for the payload)
    can_id, payload = struct.unpack('>I8s', packet_bytes)

    if can_id != EXPECTED_ID:
        print(f"Warning: Unexpected message ID: {can_id:08X}")
        return None

    # Unpack the 8-byte payload into four 16-bit Little-Endian integers
    # "<HHHH" = Little-Endian, 4x Unsigned Shorts (16-bit)
    val1, val2, val3, val4 = struct.unpack('<HHHH', payload)
    
    # --- Apply decoding logic based on our hypotheses ---
    # (You will need to refine these scaling factors)
    speed_rpm = val1
    throttle_or_current = val2
    odometer = val3
    
    # Hypothesis: Voltage = value / 1.68
    voltage_raw = val4
    voltage_scaled = voltage_raw / 1.68 if voltage_raw > 0 else 0

    return {
        "rpm": speed_rpm,
        "throttle_current": throttle_or_current,
        "odometer": odometer,
        "voltage_raw": voltage_raw,
        "voltage_scaled": f"{voltage_scaled:.2f} V"
    }
# show all the data spaced out one char at a time, two chars at a time
# show a signifier on right if the data is different from the previous data

def parse_hex_file(filename):
    """Parse a text file containing hex data."""
    try:
        with open(filename, 'r') as file:
            for line_num, line in enumerate(file, 1):
                line = line.strip()
                if not line or line.startswith('#'):  # Skip empty lines and comments
                    continue
                
                # Remove any whitespace and convert to bytes
                hex_string = ''.join(line.split())

                # split by 18F88006, not sure if this will be accurate every time
                all_signals = hex_string.split("18F88006")

                # compares each whole signal to previous signal
                prev_signal = ""
                for signal_num, signal in enumerate(all_signals, 1):
                    different_than_previous = len(prev_signal) > 0 and signal != prev_signal
                    diff_signifier = " <===" if different_than_previous else " "
                    print(f"{signal_num:3d} | ", signal, diff_signifier)
                    prev_signal = signal

                all_signals_by_packets = {}
                # create excel file
                import openpyxl
                from openpyxl.styles import PatternFill
                
                # Create a new workbook and select the active sheet
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Packet Data"
                
                # Create green fill for highlighting changed packets
                green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                
                # Add headers
                headers = ["Signal #"] + [f"Packet {i}" for i in range(14)]  # 0-13 packets
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                current_excel_row = 2  # Start data from row 2
                
                for signal_num, signal in enumerate(all_signals, 1):
                    # split by 2 chars at a time
                    packets = []
                    prev_sig_num = signal_num - 1
                    if len(signal) == 0:
                        continue
                    # start new row in excel
                    for i in range(0, len(signal), 2):
                        packets.append(signal[i:i+2])

                    for p in range(0, len(packets)):
                        current_packet_val = packets[p]
                        prev_packet_val = None
                        if prev_sig_num in all_signals_by_packets.keys() and len(all_signals_by_packets[prev_sig_num]) > p:
                            # Extract just the 2-character packet from the 4-character string
                            prev_packet_val = all_signals_by_packets[prev_sig_num][p][0:2]
                        
                        if str(current_packet_val) == str(prev_packet_val):
                            # add packet to excel file
                            packets[p] = (f"{current_packet_val}  ")
                            # Add to Excel without highlighting
                            ws.cell(row=current_excel_row, column=p+2, value=current_packet_val)
                        else:
                            # add packet to excel file, but highlight it green
                            packets[p] = (f"{current_packet_val}* ")
                            # Add to Excel with green highlighting
                            cell = ws.cell(row=current_excel_row, column=p+2, value=current_packet_val)
                            cell.fill = green_fill

                    all_signals_by_packets[signal_num] = packets
                    # Add signal number to Excel
                    ws.cell(row=current_excel_row, column=1, value=signal_num)
                    current_excel_row += 1
                    print(f"{signal_num:3d} | ", "".join(packets))

                # Save the Excel file
                excel_filename = filename.replace('.txt', '_packets.xlsx')
                wb.save(excel_filename)
                print(f"\nExcel file saved as: {excel_filename}")

                print(f"Found {len(all_signals)} complete signals")
                
                # for packet_num, packet_hex in enumerate(packets, 1):
                #     try:
                #         # Convert hex string to bytes
                #         if packet_num == 15:
                #             break;
                #         print("packet_num", packet_num)
                #         print("packet_hex", packet_hex)
                #         packet_bytes = bytes.fromhex(packet_hex)
                        
                #         # Extract the 8-byte payload (skip the 4-byte CAN ID)
                #         can_id_bytes = packet_bytes[:4]
                #         payload_bytes = packet_bytes[4:]
                        
                #         print(f"\nPacket {packet_num}:")
                #         print(f"  CAN ID: {can_id_bytes.hex().upper()}")
                #         print(f"  Payload: {payload_bytes.hex().upper()}")
                        
                #         # Parse the payload using the existing function
                #         decoded_data = parse_sabvoton_packet(packet_bytes)
                #         if decoded_data:
                #             print(f"  Parsed: {decoded_data}")
                        
                #     except ValueError as e:
                #         print(f"Packet {packet_num}: Invalid hex format - {e}")
                        
    except FileNotFoundError:
        print(f"Error: File '{filename}' not found.")
        return False
    except Exception as e:
        print(f"Error reading file: {e}")
        return False
    
    return True


def main():
    parser = argparse.ArgumentParser(description='Parse ebike data from hex text file')
    parser.add_argument('filename', help='Path to the text file containing hex data')
    
    args = parser.parse_args()
    
    print(f"Parsing file: {args.filename}")
    print("=" * 50)
    
    success = parse_hex_file(args.filename)
    
    if success:
        print("\nParsing completed successfully!")
    else:
        print("\nParsing failed!")
        sys.exit(1)


if __name__ == '__main__':
    main()

